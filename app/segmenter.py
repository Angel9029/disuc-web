import io
import uuid
import pandas as pd
import numpy as np

from datetime import date
from sqlalchemy.orm import Session
from sqlalchemy import func
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
from sklearn.decomposition import PCA
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import Lote, Producto, Pedido, Venta, DimTiempo

DOWNLOAD_STORE: dict[str, bytes] = {}


class KMeansSegmenter:

    @staticmethod
    def _get_fecha_referencia(db: Session) -> date:
        max_fecha = db.query(func.max(DimTiempo.fecha)).scalar()
        if max_fecha is None:
            return date.today()
        return max_fecha

    @staticmethod
    def _load_lotes(db: Session) -> pd.DataFrame:
        rows = (
            db.query(
                Lote.lote_id,
                Lote.producto_id,
                Lote.cantidad_inicial,
                Lote.cantidad_disponible,
                Lote.fecha_vencimiento,
            )
            .all()
        )
        df = pd.DataFrame(
            rows,
            columns=[
                "Lote_ID", "Producto_ID",
                "Cantidad_Inicial", "Cantidad_Disponible",
                "Fecha_Vencimiento",
            ],
        )
        df = df.drop_duplicates(subset="Lote_ID")
        df["Fecha_Vencimiento"] = pd.to_datetime(
            df["Fecha_Vencimiento"], errors="coerce"
        )
        df = df.dropna(subset=["Fecha_Vencimiento"])
        df["Cantidad_Inicial"] = df["Cantidad_Inicial"].clip(lower=0)
        df["Cantidad_Disponible"] = df["Cantidad_Disponible"].clip(lower=0)
        return df

    @staticmethod
    def _build_lote_vigente(
        lote_limpio: pd.DataFrame, fecha_ref: date
    ) -> pd.DataFrame:
        df = lote_limpio.copy()
        df["Dias_Para_Vencer"] = (
            df["Fecha_Vencimiento"] - pd.Timestamp(fecha_ref)
        ).dt.days
        vigente = df[
            (df["Dias_Para_Vencer"] > 0) & (df["Cantidad_Disponible"] > 0)
        ].copy()
        return vigente

    @staticmethod
    def _calc_demanda_mensual(
        db: Session, lote_limpio: pd.DataFrame
    ) -> pd.DataFrame:
        lote_ids_hist = lote_limpio["Lote_ID"].tolist()
        if not lote_ids_hist:
            return pd.DataFrame(columns=["Producto_ID", "Demanda_Mensual"])
        batch_size = 500
        resultados = []
        for i in range(0, len(lote_ids_hist), batch_size):
            batch = lote_ids_hist[i : i + batch_size]
            filas = (
                db.query(
                    Lote.producto_id,
                    DimTiempo.anio,
                    DimTiempo.mes,
                    func.sum(Pedido.cantidad_solicitada).label("total"),
                )
                .select_from(Pedido)
                .join(Venta, Pedido.venta_id == Venta.venta_id)
                .join(
                    DimTiempo,
                    Venta.fecha_id == DimTiempo.fecha_id,
                )
                .join(Lote, Pedido.lote_id == Lote.lote_id)
                .filter(Lote.lote_id.in_(batch))
                .group_by(Lote.producto_id, DimTiempo.anio, DimTiempo.mes)
                .all()
            )
            for f in filas:
                resultados.append(
                    (f.producto_id, int(f.anio), int(f.mes), float(f.total or 0))
                )
        if not resultados:
            return pd.DataFrame(columns=["Producto_ID", "Demanda_Mensual"])
        df_demanda = pd.DataFrame(
            resultados, columns=["Producto_ID", "Anio", "Mes", "Cantidad_Solicitada"]
        )
        demanda_prom = (
            df_demanda.groupby("Producto_ID")["Cantidad_Solicitada"]
            .mean()
            .reset_index()
            .rename(columns={"Cantidad_Solicitada": "Demanda_Mensual"})
        )
        return demanda_prom

    @staticmethod
    def _build_features(
        lote_vigente: pd.DataFrame,
        demanda_prom: pd.DataFrame,
        db: Session,
    ) -> pd.DataFrame:
        df = lote_vigente.merge(
            demanda_prom, on="Producto_ID", how="left"
        )
        df["Demanda_Mensual"] = df["Demanda_Mensual"].fillna(0)
        df["Rotacion"] = np.where(
            df["Cantidad_Disponible"] > 0,
            df["Demanda_Mensual"] / df["Cantidad_Disponible"],
            0.0,
        )
        precios = {
            r.producto_id: r.precio_unitario
            for r in db.query(Producto.producto_id, Producto.precio_unitario).all()
        }
        df["Precio_Unitario"] = df["Producto_ID"].map(
            lambda pid: precios.get(pid, 0)
        )
        df["Precio_Unitario"] = df["Precio_Unitario"].fillna(0)
        df["Valor_Lote"] = df["Cantidad_Disponible"] * df["Precio_Unitario"]
        return df

    @staticmethod
    def _evaluate_k(X_scaled: np.ndarray) -> dict:
        resultados = {}
        for k in range(2, 8):
            km = KMeans(n_clusters=k, random_state=42, n_init=15)
            labels = km.fit_predict(X_scaled)
            resultados[k] = {
                "inertia": float(km.inertia_),
                "silhouette": float(silhouette_score(X_scaled, labels)),
            }
        return resultados

    @staticmethod
    def _run_pca(X_scaled: np.ndarray) -> tuple:
        pca = PCA(n_components=2, random_state=42)
        X_pca = pca.fit_transform(X_scaled)
        var_exp = float(pca.explained_variance_ratio_.sum())
        return X_pca, var_exp

    @staticmethod
    def _assign_risk_labels(
        df: pd.DataFrame, kmeans: KMeans
    ) -> pd.DataFrame:
        centros = kmeans.cluster_centers_
        indice_riesgo = -centros[:, 0] + centros[:, 1] - centros[:, 2]
        orden = np.argsort(-indice_riesgo)
        mapeo = {
            orden[0]: "Riesgo Alto",
            orden[1]: "Riesgo Medio",
            orden[2]: "Riesgo Bajo",
        }
        df["Nivel_Riesgo"] = df["Cluster_ID"].map(mapeo)
        return df, mapeo, indice_riesgo

    @staticmethod
    def _sensitivity_analysis(centros: np.ndarray) -> list[dict]:
        signos = np.array([-1, 1, -1])
        escenarios = {
            "Peso igual entre las 3 variables": np.array([1, 1, 1]),
            "Doble peso al Valor del Lote": np.array([1, 2, 1]),
            "Doble peso a Dias para Vencer": np.array([2, 1, 1]),
            "Doble peso a Rotacion": np.array([1, 1, 2]),
        }
        resultados = []
        for nombre, pesos in escenarios.items():
            indice = (centros * signos * pesos).sum(axis=1)
            orden = list(np.argsort(-indice))
            resultados.append({"escenario": nombre, "orden": orden})
        return resultados

    @staticmethod
    def _generate_excel(
        df: pd.DataFrame, mapeo: dict
    ) -> tuple[bytes, list[dict]]:
        col_map = {
            "Lote_ID": "Lote_ID",
            "Producto_ID": "Producto_ID",
            "Dias_Para_Vencer": "Dias_Para_Vencer",
            "Cantidad_Disponible": "Cantidad_Disponible",
            "Precio_Unitario": "Precio_Unitario",
            "Valor_Lote": "Valor_Lote",
            "Rotacion": "Rotacion",
            "Nivel_Riesgo": "Nivel_Riesgo",
        }
        detalle = df[list(col_map.keys())].copy()
        resumen_data = (
            detalle.groupby("Nivel_Riesgo")
            .agg(
                Cantidad_Lotes=("Lote_ID", "count"),
                Dias_Promedio=("Dias_Para_Vencer", "mean"),
                Stock_Total=("Cantidad_Disponible", "sum"),
                Rotacion_Promedio=("Rotacion", "mean"),
                Valor_Total=("Valor_Lote", "sum"),
            )
            .reindex(["Riesgo Alto", "Riesgo Medio", "Riesgo Bajo"])
            .reset_index()
        )
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            resumen_data.to_excel(
                writer, sheet_name="Resumen_Riesgo", index=False
            )
            detalle.to_excel(
                writer, sheet_name="Detalle_Lotes_Riesgo", index=False
            )
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        fill_header = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        font_header = Font(
            name="Calibri", size=11, bold=True, color="FFFFFF"
        )
        fill_alto = PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        )
        fill_medio = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )
        fill_bajo = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )
        font_alto = Font(
            name="Calibri", size=11, color="C00000", bold=True
        )
        font_medio = Font(
            name="Calibri", size=11, color="7F6000", bold=True
        )
        font_bajo = Font(
            name="Calibri", size=11, color="375623", bold=True
        )
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        col_formats: dict[str, str] = {
            "Cantidad_Disponible": '#,##0',
            "Precio_Unitario": '"S/" #,##0.00',
            "Valor_Lote": '"S/" #,##0.00',
            "Rotacion": "0.000",
            "Dias_Para_Vencer": '#,##0',
        }
        fill_map = {
            "Riesgo Alto": (fill_alto, font_alto),
            "Riesgo Medio": (fill_medio, font_medio),
            "Riesgo Bajo": (fill_bajo, font_bajo),
        }
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            ws.row_dimensions[1].height = 28
            header_cols = [
                cell.value for cell in ws[1]
            ]
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 19
                for col_idx, cell in enumerate(ws[row], start=1):
                    cell.border = thin_border
                    cell.font = Font(name="Calibri", size=11)
                    if ws.title == "Detalle_Lotes_Riesgo":
                        col_name = header_cols[col_idx - 1] if col_idx <= len(header_cols) else ""
                        fmt = col_formats.get(str(col_name))
                        if fmt:
                            cell.number_format = fmt
                            cell.alignment = Alignment(horizontal="right")
                        if str(col_name) == "Nivel_Riesgo":
                            cell.alignment = Alignment(horizontal="center")
                            if cell.value in fill_map:
                                f, fn = fill_map[cell.value]
                                cell.fill = f
                                cell.font = fn
                    elif ws.title == "Resumen_Riesgo":
                        if col_idx == 1:
                            cell.alignment = Alignment(horizontal="center")
                            cell.font = Font(name="Calibri", size=11, bold=True)
                            if cell.value in fill_map:
                                f, fn = fill_map[cell.value]
                                cell.fill = f
                                cell.font = fn
                        elif col_idx == 2:
                            cell.number_format = "#,##0"
                            cell.alignment = Alignment(horizontal="right")
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for c in col_cells:
                    if c.value is not None:
                        val_str = str(c.value)
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        buf2 = io.BytesIO()
        wb.save(buf2)
        buf2.seek(0)
        resumen_list = resumen_data.to_dict(orient="records")
        return buf2.getvalue(), resumen_list

    @staticmethod
    def _clean(obj):
        if isinstance(obj, dict):
            return {k: KMeansSegmenter._clean(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [KMeansSegmenter._clean(v) for v in obj]
        elif isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return obj

    @staticmethod
    def segment(db: Session) -> dict:
        fecha_ref = KMeansSegmenter._get_fecha_referencia(db)
        lote_limpio = KMeansSegmenter._load_lotes(db)
        if lote_limpio.empty:
            raise ValueError("No se encontraron lotes en la base de datos")
        lote_vigente = KMeansSegmenter._build_lote_vigente(
            lote_limpio, fecha_ref
        )
        if lote_vigente.empty:
            raise ValueError("No se encontraron lotes vigentes")
        demanda_prom = KMeansSegmenter._calc_demanda_mensual(
            db, lote_limpio
        )
        lote_modelo = KMeansSegmenter._build_features(
            lote_vigente, demanda_prom, db
        )
        features = ["Dias_Para_Vencer", "Valor_Lote", "Rotacion"]
        X = lote_modelo[features].values
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        eval_k = KMeansSegmenter._evaluate_k(X_scaled)
        X_pca, var_exp = KMeansSegmenter._run_pca(X_scaled)
        kmeans = KMeans(n_clusters=3, random_state=42, n_init=15)
        lote_modelo["Cluster_ID"] = kmeans.fit_predict(X_scaled)
        sil_final = float(silhouette_score(X_scaled, kmeans.labels_))
        lote_modelo, mapeo, indices_riesgo = (
            KMeansSegmenter._assign_risk_labels(lote_modelo, kmeans)
        )
        sensibilidad = KMeansSegmenter._sensitivity_analysis(
            kmeans.cluster_centers_
        )
        excel_bytes, resumen_riesgo = KMeansSegmenter._generate_excel(
            lote_modelo, mapeo
        )
        token = uuid.uuid4().hex
        DOWNLOAD_STORE[token] = excel_bytes
        summary = (
            lote_modelo.groupby("Nivel_Riesgo")
            .agg(Total=("Lote_ID", "count"))
            .to_dict()["Total"]
        )
        producto_map = {
            r.producto_id: r.nombre
            for r in db.query(Producto).all()
        }
        datos = []
        for _, row in lote_modelo.iterrows():
            datos.append({
                "Lote_ID": int(row["Lote_ID"]),
                "Producto_ID": int(row["Producto_ID"]),
                "Producto": producto_map.get(
                    int(row["Producto_ID"]), f"ID {int(row['Producto_ID'])}"
                ),
                "Dias_Para_Vencer": int(row["Dias_Para_Vencer"]),
                "Cantidad_Disponible": int(row["Cantidad_Disponible"]),
                "Valor_Lote": round(float(row["Valor_Lote"]), 2),
                "Rotacion": round(float(row["Rotacion"]), 4),
                "Cluster_ID": int(row["Cluster_ID"]),
                "Nivel_Riesgo": row["Nivel_Riesgo"],
            })
        centroides = []
        for orig_idx in range(3):
            riesgo_label = mapeo.get(orig_idx, f"Cluster {orig_idx}")
            centroides.append({
                "cluster": int(orig_idx),
                "riesgo": riesgo_label,
                "Dias_Para_Vencer": round(
                    float(scaler.inverse_transform(
                        kmeans.cluster_centers_
                    )[orig_idx, 0]), 2
                ),
                "Valor_Lote": round(
                    float(scaler.inverse_transform(
                        kmeans.cluster_centers_
                    )[orig_idx, 1]), 2
                ),
                "Rotacion": round(
                    float(scaler.inverse_transform(
                        kmeans.cluster_centers_
                    )[orig_idx, 2]), 4
                ),
            })
        return KMeansSegmenter._clean({
            "fecha_referencia": fecha_ref.isoformat(),
            "total_lotes": len(lote_modelo),
            "resumen": {
                "Riesgo Alto": summary.get("Riesgo Alto", 0),
                "Riesgo Medio": summary.get("Riesgo Medio", 0),
                "Riesgo Bajo": summary.get("Riesgo Bajo", 0),
            },
            "evaluacion_k": eval_k,
            "silhouette_final": sil_final,
            "varianza_pca": var_exp,
            "datos": datos,
            "centroides": centroides,
            "indices_riesgo": {
                str(int(k)): round(float(v), 4)
                for k, v in enumerate(indices_riesgo)
            },
            "sensibilidad": sensibilidad,
            "resumen_riesgo": resumen_riesgo,
            "download_token": token,
        })
